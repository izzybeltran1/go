from openpyxl import Workbook
from openpyxl import load_workbook
import sys, json

START_ROW = 3
END_ROW = 10000
DATA_COLS = ['A', 'B', 'C', 'D', 'E']
file_name = '2_bad.xlsx'
wb = load_workbook(file_name)
ws = wb.active

#Читаем строки таблицы
result = []
for row in range(START_ROW, END_ROW):
    # Если ячейка пуста - конец
    if ws[f"A{row}"].value == None:
        break

    result.append( [int(ws[f"{col}{row}"].value) for col in DATA_COLS] )

#Возврат результата в PHP
for i in result:
    print(json.dumps(i))
